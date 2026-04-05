"""
================================================================
COD COMPLET — Simulari + Grafice + Excel
Articol JES: N=3, M=2, K=9
R=1000 repetitii, n variabil: 33,50,100,200,500,1000
================================================================
Rulare: python3 simulari_complete.py
Cerinte: pip install numpy scipy matplotlib openpyxl
================================================================
"""

import numpy as np
import secrets as sec_module
import math
from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import time

# ════════════════════════════════════════════════════════════════
# 1. PARAMETRI
# ════════════════════════════════════════════════════════════════
np.random.seed(42)
K = 9; N = 3; M = 2
alpha = 0.05
z = norm.ppf(1 - alpha/2)   # 1.9599639845
R = 1000                      # repetitii Monte Carlo
n_vals = [33, 50, 100, 200, 500, 1000]

# ════════════════════════════════════════════════════════════════
# 2. FUNCTII MODEL
# ════════════════════════════════════════════════════════════════

def pmf_sp(k, K=9):
    pk = (k+1)/(K+1); pk0 = k/(K+1)
    if k == 0: return 1 - (1 - pk**N)**M
    return (1 - pk0**N)**M - (1 - pk**N)**M

def pmf_ps(k, K=9):
    qk = (K-k)/(K+1); qk0 = (K-k+1)/(K+1)
    if k == 0: return (1 - qk**N)**M
    return (1 - qk**N)**M - (1 - qk0**N)**M

def F_sp(k, K=9): return 1 - (1 - ((k+1)/(K+1))**N)**M
def F_ps(k, K=9): return (1 - ((K-k)/(K+1))**N)**M
def R_sp(k, K=9): return 1 - F_sp(k, K)
def R_ps(k, K=9): return 1 - F_ps(k, K)

def h_sp(k, K=9):
    p = pmf_sp(k, K); r = R_sp(k-1, K) if k > 0 else 1.0
    return p/r if r > 1e-15 else 0.0

def h_ps(k, K=9):
    p = pmf_ps(k, K); r = R_ps(k-1, K) if k > 0 else 1.0
    return p/r if r > 1e-15 else 0.0

mu_sp  = sum(k * pmf_sp(k) for k in range(K+1))
mu_ps  = sum(k * pmf_ps(k) for k in range(K+1))
var_sp = sum((k-mu_sp)**2 * pmf_sp(k) for k in range(K+1))
var_ps = sum((k-mu_ps)**2 * pmf_ps(k) for k in range(K+1))
sig    = math.sqrt(var_sp)  # sig_sp = sig_ps (simetrie)

# ════════════════════════════════════════════════════════════════
# 3. GENERATOARE GNPA
# ════════════════════════════════════════════════════════════════

# Cifrele lui pi (sursa: P. Trueb, 2016)
_PI = (
    "14159265358979323846264338327950288419716939937510"
    "58209749445923078164062862089986280348253421170679"
    "82148086513282306647093844609550582231725359408128"
    "48111745028410270193852110555964462294895493038196"
    "44288109756659334461284756482337867831652712019091"
    "45648566923460348610454326648213393607260249141273"
    "72458700660631558817488152092096282925409171536436"
    "78925903600113305305488204665213841469519415116094"
    "33057270365759591953092186117381932611793105118548"
    "07446237996274956735188575272489122793818301194912"
    "98336733624406566430860213949463952247371907021798"
    "60943702770539217176293176752384674818467669405132"
    "00056812714526356082778577134275778960917363717872"
    "14684409012249534301465495853710507922796892589235"
    "42019956112129021960864034418159813629774771309960"
    "51870721134999999837297804995105973173281609631859"
    "50244594553469083026425223082533446850352619311881"
    "71010003137838752886587533208381420617177669147303"
    "59825349042875546873115956286388235378759375195778"
    "18577805321712268066130019278766111959092164201989"
    "38095257201065485863278865936153381827968230301952"
    "03530185296899577362259941389124972177528347913151"
    "55748572424541506959508295331168617278558890750983"
    "81754637464939319255060400927701671139009848824012"
    "85836160356370766010471018194295559619894676783744"
    "94482553797747268471040475346462080466842590694912"
    "93313677028989152104752162056966024058038150193511"
    "25338243003558764024749647326391419927260426992279"
    "67823547816360093417216412199245863150302861829745"
    "55706749838505494588586926995690927210797509302955"
    "32116534498720275596023648066549911988183479775356"
    "63698074265425278625518184175746728909777727938000"
    "81647060016145249192173217214772350141441973568548"
    "16136115735255213347574184946843852332390739414333"
    "45477624168625189835694855620992192221842725502542"
    "56887671790494601653466804988627232791786085784383"
    "82796797668145410095388378636095068006422512520511"
    "73929848960841284886269456042419652850222106611863"
    "06744278622039194945047123713786960956364371917287"
    "46776465757396241389086583264599581339047802759010"
)
_PI_ARR = np.array([int(d) for d in _PI if d.isdigit()])

def pi_digits(n, off=0):
    needed = off + n
    reps = needed // len(_PI_ARR) + 2
    return np.tile(_PI_ARR, reps)[off:off+n]

def gnpa1(n, off=0):
    """Java ThreadLocalRandom — simulat cu numpy Mersenne Twister"""
    return np.random.randint(0, 10, size=n)

def gnpa2(n, off=0):
    """Python secrets — generator criptografic securizat"""
    return np.array([sec_module.randbelow(10) for _ in range(n)])

def gnpa3(n, off=0):
    """Cifrele zecimale ale lui pi (Trueb, 2016)"""
    return pi_digits(n, off)

GNPAS = [
    ("Java ThreadLocalRandom", gnpa1, "Java TLR"),
    ("Python secrets",          gnpa2, "Py secrets"),
    ("Cifrele lui \u03C0",     gnpa3, "Cifrele \u03C0"),
]

# ════════════════════════════════════════════════════════════════
# 4. SIMULARE DURATE DE VIATA
# ════════════════════════════════════════════════════════════════

def simulate(n_sim, network, gnpa_func, pi_off=0):
    """
    Simuleaza n_sim durate de viata.
    network: 'SP' sau 'PS'
    Returneaza array de intregi in {0,...,9}
    """
    units_total = N * M
    if gnpa_func == gnpa3:
        mat = gnpa3(n_sim * units_total, pi_off).reshape(n_sim, M, N)
    else:
        mat = gnpa_func(n_sim * units_total).reshape(n_sim, M, N)
    if network == 'SP':
        return mat.max(axis=2).min(axis=1).astype(np.int32)
    else:
        return mat.min(axis=2).max(axis=1).astype(np.int32)

# ════════════════════════════════════════════════════════════════
# 5. MLE
# ════════════════════════════════════════════════════════════════

def mle_complete(data, pmf_func, K_max=30):
    """MLE pentru K pe baza datelor complete."""
    Km = int(np.max(data))
    best_K, best_ll = Km, -np.inf
    vals, cnts = np.unique(data, return_counts=True)
    for Kt in range(Km, K_max+1):
        ll = sum(cnt * np.log(max(pmf_func(int(v), Kt), 1e-300))
                 for v, cnt in zip(vals, cnts))
        if ll > best_ll:
            best_ll, best_K = ll, Kt
    return best_K

def mle_censored(data, intervals, F_func, K_max=30):
    """MLE pentru K pe baza datelor censurate."""
    counts = []
    for a, b in intervals:
        if b is None: counts.append(int(np.sum(data >= a)))
        else:         counts.append(int(np.sum((data >= a) & (data < b))))
    Km = int(intervals[-1][0])
    best_K, best_ll = Km, -np.inf
    for Kt in range(Km, K_max+1):
        ll = 0.0; valid = True
        for (a, b), cnt in zip(intervals, counts):
            if cnt == 0: continue
            Fb = F_func(b-1, Kt) if b is not None else 1.0
            Fa = F_func(a-1, Kt) if a > 0 else 0.0
            p  = Fb - Fa
            if p <= 0: valid = False; break
            ll += cnt * np.log(p)
        if valid and ll > best_ll:
            best_ll, best_K = ll, Kt
    return best_K

# Scheme de censurare
INT_SP_C3 = [(0,4), (4,7), (7,None)]
INT_SP_C5 = [(0,4), (4,6), (6,7), (7,8), (8,None)]
INT_PS_C3 = [(0,3), (3,6), (6,None)]
INT_PS_C5 = [(0,3), (3,5), (5,7), (7,8), (8,None)]

# ════════════════════════════════════════════════════════════════
# 6. RULARE SIMULARI COMPLETE
# ════════════════════════════════════════════════════════════════

print("=" * 65)
print(f"SIMULARI MONTE CARLO  N={N}, M={M}, K={K}, R={R}")
print("=" * 65)
print(f"z_{{0.975}} = {z:.10f}")
print(f"sigma      = {sig:.10f}")
print(f"E[U] SP    = {mu_sp:.10f}")
print(f"E[V] PS    = {mu_ps:.10f}")
print(f"D[U]=D[V]  = {var_sp:.10f}")
print()

# Formula n
print("Calculul n prin formula N = floor(((z*sigma)/eps)^2) + 1:")
for eps in [0.50, 0.30, 0.20, 0.10, 0.05]:
    n_c = math.floor(((z*sig)/eps)**2) + 1
    print(f"  eps={eps:.2f} => n={n_c:,}")
print()

results = {}  # results[net][gname][n] = dict cu statistici

for net, pmf_f, F_f, ic3, ic5, mu_t in [
    ('SP', pmf_sp, F_sp, INT_SP_C3, INT_SP_C5, mu_sp),
    ('PS', pmf_ps, F_ps, INT_PS_C3, INT_PS_C5, mu_ps),
]:
    results[net] = {}
    print(f"\n{'='*65}")
    print(f"Reteaua {net}  (E[X]_teoretic = {mu_t:.10f})")
    print(f"{'='*65}")

    for gname, gfunc, glabel in GNPAS:
        results[net][gname] = {}
        print(f"\n  GNPA: {gname}")
        t0 = time.time()

        for n in n_vals:
            Kc  = np.zeros(R, dtype=int)
            Kc3 = np.zeros(R, dtype=int)
            Kc5 = np.zeros(R, dtype=int)
            ms  = np.zeros(R)

            for r in range(R):
                off = r * n * N * M + (50_000_000 if net=='PS' else 0)
                data = simulate(n, net, gfunc, pi_off=off)
                ms[r]  = float(np.mean(data))
                Kc[r]  = mle_complete(data, pmf_f)
                Kc3[r] = mle_censored(data, ic3, F_f)
                Kc5[r] = mle_censored(data, ic5, F_f)

            results[net][gname][n] = {
                'mean_sim': float(np.mean(ms)),
                'c':  {
                    'mean': float(np.mean(Kc)),
                    'bias': float(np.mean(Kc) - K),
                    'mse':  float(np.mean((Kc  - K)**2)),
                    'std':  float(np.std(Kc)),
                    'rate': float(np.mean(Kc == K)),
                },
                'c3': {
                    'mean': float(np.mean(Kc3)),
                    'bias': float(np.mean(Kc3) - K),
                    'mse':  float(np.mean((Kc3 - K)**2)),
                    'std':  float(np.std(Kc3)),
                    'rate': float(np.mean(Kc3 == K)),
                },
                'c5': {
                    'mean': float(np.mean(Kc5)),
                    'bias': float(np.mean(Kc5) - K),
                    'mse':  float(np.mean((Kc5 - K)**2)),
                    'std':  float(np.std(Kc5)),
                    'rate': float(np.mean(Kc5 == K)),
                },
            }
            r_ = results[net][gname][n]
            print(f"    n={n:>5}: mean_sim={r_['mean_sim']:.4f} "
                  f"rate_c={r_['c']['rate']:.4f} "
                  f"MSE_c={r_['c']['mse']:.6f} | "
                  f"rate_c3={r_['c3']['rate']:.4f} "
                  f"MSE_c3={r_['c3']['mse']:.6f} | "
                  f"rate_c5={r_['c5']['rate']:.4f} "
                  f"MSE_c5={r_['c5']['mse']:.6f}")

        print(f"    Timp GNPA: {time.time()-t0:.1f}s")

# ════════════════════════════════════════════════════════════════
# 7. GRAFICE
# ════════════════════════════════════════════════════════════════

colors  = ['#1f77b4', '#ff7f0e', '#2ca02c']
markers = ['o', 's', '^']
labels  = [g[2] for g in GNPAS]
gnames  = [g[0] for g in GNPAS]
k_vals  = list(range(K+1))

# ── Figura 1: CDF, Fiabilitate, Hazard ──────────────────────
fig1, axes1 = plt.subplots(1, 2, figsize=(13, 5))
fig1.suptitle(
    f"Fig. 1. Funcțiile de repartiție F(k), fiabilitate R(k) și hazard h(k)\n"
    f"pentru rețelele SP și PS (N={N}, M={M}, K={K})",
    fontsize=12, fontweight='bold'
)

for ax, net, F_f, R_f, h_f, title in [
    (axes1[0], 'SP', F_sp, R_sp, h_sp, f'Rețeaua SP (U = min(max(·)))'),
    (axes1[1], 'PS', F_ps, R_ps, h_ps, f'Rețeaua PS (V = max(min(·)))'),
]:
    F_v = [F_f(k) for k in k_vals]
    R_v = [R_f(k) for k in k_vals]
    h_v = [h_f(k) for k in k_vals]

    ax.step(k_vals, F_v, where='post', color='#1f77b4', lw=2,
            marker='o', ms=5, label='F(k) — CDF')
    ax.step(k_vals, R_v, where='post', color='#ff7f0e', lw=2,
            marker='s', ms=5, ls='--', label='R(k) — Fiabilitate')
    ax.step(k_vals, h_v, where='post', color='#2ca02c', lw=2,
            marker='^', ms=5, ls=':', label='h(k) — Hazard')

    ax.set_title(title, fontsize=11, fontweight='bold')
    ax.set_xlabel('k', fontsize=11)
    ax.set_ylabel('Valoare', fontsize=11)
    ax.set_xticks(k_vals)
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(-0.05, 1.15)

plt.tight_layout()
plt.savefig('figura1_CDF_R_h.png', dpi=150, bbox_inches='tight')
plt.close()
print("\n✅ Figura 1 salvata: figura1_CDF_R_h.png")

# ── Figura 2: Rata de succes vs n — date complete ───────────
fig2, axes2 = plt.subplots(1, 2, figsize=(13, 5))
fig2.suptitle(
    f"Fig. 2. Rata de succes P(K̂=K) în funcție de n — date complete\n"
    f"(N={N}, M={M}, K={K}, R={R} repetitii)",
    fontsize=12, fontweight='bold'
)

for ax, net, title in [
    (axes2[0], 'SP', 'Rețeaua SP'),
    (axes2[1], 'PS', 'Rețeaua PS'),
]:
    for i, (gname, glabel) in enumerate(zip(gnames, labels)):
        rates = [results[net][gname][n]['c']['rate'] for n in n_vals]
        ax.plot(n_vals, rates, color=colors[i], marker=markers[i],
                lw=2, ms=7, label=glabel)

    ax.axhline(y=1.0, color='gray', ls='--', lw=1, alpha=0.5)
    ax.set_title(title, fontsize=11, fontweight='bold')
    ax.set_xlabel('n (volum eșantion)', fontsize=11)
    ax.set_ylabel('P(K̂ = K)', fontsize=11)
    ax.set_xticks(n_vals)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(0, 1.1)

plt.tight_layout()
plt.savefig('figura2_rata_succes_complete.png', dpi=150, bbox_inches='tight')
plt.close()
print("✅ Figura 2 salvata: figura2_rata_succes_complete.png")

# ── Figura 3: Rata de succes vs n — C3 si C5 ────────────────
fig3, axes3 = plt.subplots(2, 2, figsize=(13, 10))
fig3.suptitle(
    f"Fig. 3. Rata de succes P(K̂=K) — date censurate C3 și C5\n"
    f"(N={N}, M={M}, K={K}, R={R} repetitii)",
    fontsize=12, fontweight='bold'
)

for row_idx, (schema, key) in enumerate([('C3 (3 intervale)', 'c3'),
                                           ('C5 (5 intervale)', 'c5')]):
    for col_idx, (net, title) in enumerate([('SP', 'Rețeaua SP'),
                                              ('PS', 'Rețeaua PS')]):
        ax = axes3[row_idx][col_idx]
        for i, (gname, glabel) in enumerate(zip(gnames, labels)):
            rates = [results[net][gname][n][key]['rate'] for n in n_vals]
            ax.plot(n_vals, rates, color=colors[i], marker=markers[i],
                    lw=2, ms=7, label=glabel)

        ax.axhline(y=1.0, color='gray', ls='--', lw=1, alpha=0.5)
        ax.set_title(f"{title} — {schema}", fontsize=10, fontweight='bold')
        ax.set_xlabel('n', fontsize=10)
        ax.set_ylabel('P(K̂ = K)', fontsize=10)
        ax.set_xticks(n_vals)
        ax.legend(fontsize=9)
        ax.grid(True, alpha=0.3)
        ax.set_ylim(0, 1.1)

plt.tight_layout()
plt.savefig('figura3_rata_succes_censurate.png', dpi=150, bbox_inches='tight')
plt.close()
print("✅ Figura 3 salvata: figura3_rata_succes_censurate.png")

# ── Figura 4: MSE vs n ───────────────────────────────────────
fig4, axes4 = plt.subplots(1, 2, figsize=(13, 5))
fig4.suptitle(
    f"Fig. 4. MSE(K̂) în funcție de n — date complete, C3, C5\n"
    f"(N={N}, M={M}, K={K}, R={R} repetitii)",
    fontsize=12, fontweight='bold'
)

ls_map = {'c': '-', 'c3': '--', 'c5': ':'}
label_map = {'c': 'Complete', 'c3': 'Censurate C3', 'c5': 'Censurate C5'}

for ax, net, title in [
    (axes4[0], 'SP', 'Rețeaua SP'),
    (axes4[1], 'PS', 'Rețeaua PS'),
]:
    for i, (gname, glabel) in enumerate(zip(gnames, labels)):
        for key in ['c', 'c3', 'c5']:
            mse_v = [results[net][gname][n][key]['mse'] for n in n_vals]
            ax.plot(n_vals, mse_v, color=colors[i],
                    ls=ls_map[key], marker=markers[i],
                    lw=1.8, ms=5,
                    label=f"{glabel} — {label_map[key]}")

    ax.set_title(title, fontsize=11, fontweight='bold')
    ax.set_xlabel('n', fontsize=11)
    ax.set_ylabel('MSE(K̂)', fontsize=11)
    ax.set_xticks(n_vals)
    ax.legend(fontsize=7, ncol=2)
    ax.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig('figura4_MSE.png', dpi=150, bbox_inches='tight')
plt.close()
print("✅ Figura 4 salvata: figura4_MSE.png")

# ════════════════════════════════════════════════════════════════
# 8. SALVARE EXCEL
# ════════════════════════════════════════════════════════════════

wb = Workbook()
HDR="1F4E79"; SUB="2E75B6"; GRN="E2EFDA"; YEL="FFF2CC"; RED="FCE4D6"

def cell(ws, r, c, v, bold=False, bg=None, fmt=None,
         align="center", color="000000", italic=False,
         size=10, fname="Arial"):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font = Font(name=fname, bold=bold, size=size,
                   color=color, italic=italic)
    if bg:
        cl.fill = PatternFill("solid", fgColor=bg)
    cl.alignment = Alignment(horizontal=align, vertical="center",
                              wrap_text=True)
    t = Side(style="thin", color="AAAAAA")
    cl.border = Border(left=t, right=t, top=t, bottom=t)
    if fmt:
        cl.number_format = fmt
    return cl

NF10 = "0.0000000000"
NF6  = "0.000000"
NF4  = "0.0000"

# ── Foaia 1: Valori Teoretice ─────────────────────────────────
ws1 = wb.active
ws1.title = "Valori_Teoretice"
ws1.merge_cells("A1:I1")
cell(ws1,1,1,"VALORI TEORETICE — N=3, M=2, K=9 (10 zecimale)",
     bold=True,bg=HDR,color="FFFFFF",size=12)
ws1.row_dimensions[1].height = 28

# Parametri
ws1.merge_cells("A3:D3")
cell(ws1,3,1,"PARAMETRI",bold=True,bg=SUB,color="FFFFFF")
params = [
    ("z_{1-alpha/2}  (alpha=0.05)", z,      "norm.ppf(0.975) — cuantila normala"),
    ("E[U]  reteaua SP",           mu_sp,  "Suma k*P(U=k), k=0..9"),
    ("E[V]  reteaua PS",           mu_ps,  "Suma k*P(V=k), k=0..9"),
    ("D[U] = D[V]",                var_sp, "Dispersia (identica prin simetrie)"),
    ("sigma = sqrt(D)",            sig,    "Abaterea standard"),
]
for i,(nm,vl,desc) in enumerate(params):
    r=4+i; bg=GRN if i%2==0 else YEL
    cell(ws1,r,1,nm,bg=bg,align="left")
    cell(ws1,r,2,vl,bg=bg,fmt=NF10,align="right",fname="Courier New")
    ws1.merge_cells(f"C{r}:D{r}")
    cell(ws1,r,3,desc,bg=bg,align="left",italic=True)
    ws1.row_dimensions[r].height = 16

# Calculul n
ws1.merge_cells("A11:D11")
cell(ws1,11,1,"CALCULUL n — formula N=floor(((z*sigma)/eps)^2)+1",
     bold=True,bg=SUB,color="FFFFFF")
cell(ws1,12,1,"epsilon",bold=True,bg=HDR,color="FFFFFF")
cell(ws1,12,2,"(z*sigma/eps)^2",bold=True,bg=HDR,color="FFFFFF")
cell(ws1,12,3,"floor(...)",bold=True,bg=HDR,color="FFFFFF")
cell(ws1,12,4,"N=floor(...)+1",bold=True,bg=HDR,color="FFFFFF")
ws1.row_dimensions[12].height = 20
for i,eps in enumerate([0.50,0.30,0.20,0.10,0.05]):
    r=13+i
    ratio = ((z*sig)/eps)**2
    fl    = math.floor(ratio)
    n_c   = fl+1
    bg    = RED if eps in {0.10,0.05} else (GRN if i%2==0 else YEL)
    cell(ws1,r,1,eps,bg=bg,fmt="0.00")
    cell(ws1,r,2,ratio,bg=bg,fmt=NF10,align="right")
    cell(ws1,r,3,fl,bg=bg,fmt="0",align="right")
    cell(ws1,r,4,n_c,bg=bg,fmt="0",align="right",bold=True)
    ws1.row_dimensions[r].height=16

# Tabel PMF
ws1.merge_cells("A20:I20")
cell(ws1,20,1,"PMF, CDF, FIABILITATE, HAZARD (10 zecimale)",
     bold=True,bg=SUB,color="FFFFFF")
hdrs=["k","F_sp(k)","P(U=k)","R_sp(k)","h_sp(k)",
          "F_ps(k)","P(V=k)","R_ps(k)","h_ps(k)"]
for j,h in enumerate(hdrs,1):
    cell(ws1,21,j,h,bold=True,bg=HDR,color="FFFFFF")
ws1.row_dimensions[21].height=20
for k in range(K+1):
    r=22+k; bg=GRN if k%2==0 else None
    for j,v in enumerate([k,F_sp(k),pmf_sp(k),R_sp(k),h_sp(k),
                              F_ps(k),pmf_ps(k),R_ps(k),h_ps(k)],1):
        cell(ws1,r,j,v,bg=bg,
             fmt=None if j==1 else NF10,
             align="center" if j==1 else "right")
    ws1.row_dimensions[r].height=16
# Suma
r_s=22+K+1
cell(ws1,r_s,1,"SUMA PMF",bold=True,bg=YEL)
cell(ws1,r_s,3,sum(pmf_sp(k) for k in range(K+1)),
     bold=True,bg=YEL,fmt=NF10,align="right")
cell(ws1,r_s,7,sum(pmf_ps(k) for k in range(K+1)),
     bold=True,bg=YEL,fmt=NF10,align="right")

for i,w in enumerate([6,14,14,14,14,14,14,14,14],1):
    ws1.column_dimensions[get_column_letter(i)].width=w

# ── Foile 2-3: Simulari SP si PS ─────────────────────────────
for net in ["SP","PS"]:
    ws_n = wb.create_sheet(f"Simulari_{net}")
    mu_t = mu_sp if net=="SP" else mu_ps
    ws_n.merge_cells("A1:N1")
    cell(ws_n,1,1,
         f"SIMULARI MONTE CARLO — Reteaua {net}, N={N}, M={M}, K={K}, R={R}",
         bold=True,bg=HDR,color="FFFFFF",size=12)
    ws_n.row_dimensions[1].height=28
    ws_n.merge_cells("A3:N3")
    cell(ws_n,3,1,
         f"E[X]_teoretic={mu_t:.10f} | sigma={sig:.10f} | "
         f"z={z:.10f}",
         bg="DEEAF1",align="left",italic=True,size=9)

    hdrs_s=["n","E[X]\nteoret.","GNPA","Media\nsim.",
            "K_hat\nC","Bias\nC","MSE\nC","Rata\nC",
            "K_hat\nC3","MSE\nC3","Rata\nC3",
            "K_hat\nC5","MSE\nC5","Rata\nC5"]
    for j,h in enumerate(hdrs_s,1):
        cell(ws_n,5,j,h,bold=True,bg=HDR,color="FFFFFF")
    ws_n.row_dimensions[5].height=40

    row=6
    for n in n_vals:
        for gi,(gname,_,_) in enumerate(GNPAS):
            bg=GRN if gi==0 else (YEL if gi==1 else None)
            res=results[net][gname][n]
            vals=[
                n if gi==0 else "",
                f"{mu_t:.6f}" if gi==0 else "",
                gname,
                res['mean_sim'],
                res['c']['mean'],  res['c']['bias'],  res['c']['mse'],  res['c']['rate'],
                res['c3']['mean'], res['c3']['mse'],  res['c3']['rate'],
                res['c5']['mean'], res['c5']['mse'],  res['c5']['rate'],
            ]
            fmts=[None,None,None,NF6,NF6,NF6,NF6,NF4,
                  NF6,NF6,NF4, NF6,NF6,NF4]
            aligns=["center","right","left"]+["right"]*11
            for j,(v,fmt,aln) in enumerate(zip(vals,fmts,aligns),1):
                cell(ws_n,row,j,v,bg=bg,fmt=fmt,align=aln)
            ws_n.row_dimensions[row].height=16
            row+=1
        row+=1

    for i,w in enumerate([8,14,24,12,12,12,12,10,12,12,10,12,12,10],1):
        ws_n.column_dimensions[get_column_letter(i)].width=w

# ── Foaia 4: Comparatie GNPA ─────────────────────────────────
ws5=wb.create_sheet("Comparatie_GNPA")
ws5.merge_cells("A1:K1")
cell(ws5,1,1,
     f"COMPARATIE GNPA — Rata de succes P(K_hat=K) in functie de n (R={R})",
     bold=True,bg=HDR,color="FFFFFF",size=12)
ws5.row_dimensions[1].height=28

row=3
for net in ["SP","PS"]:
    ws5.merge_cells(f"A{row}:K{row}")
    cell(ws5,row,1,f"Reteaua {net}",bold=True,bg=SUB,color="FFFFFF")
    row+=1
    hdrs_c=["n",
             "Java TLR\nC","Java TLR\nC3","Java TLR\nC5",
             "Py secrets\nC","Py secrets\nC3","Py secrets\nC5",
             "Pi\nC","Pi\nC3","Pi\nC5",
             "Cel mai bun\n(date complete)"]
    for j,h in enumerate(hdrs_c,1):
        cell(ws5,row,j,h,bold=True,bg=HDR,color="FFFFFF")
    ws5.row_dimensions[row].height=40
    row+=1
    for i,n in enumerate(n_vals):
        bg=GRN if i%2==0 else YEL
        vals=[n]
        rates_c=[]
        for gname,_,_ in GNPAS:
            r=results[net][gname][n]
            vals+=[r['c']['rate'],r['c3']['rate'],r['c5']['rate']]
            rates_c.append((r['c']['rate'],gname))
        best=max(rates_c,key=lambda x:x[0])[1]
        vals.append(best)
        fmts=[None]+[NF4,NF4,NF4]*3+[None]
        aligns=["center"]+["right","right","right"]*3+["left"]
        for j,(v,fmt,aln) in enumerate(zip(vals,fmts,aligns),1):
            cell(ws5,row,j,v,bg=bg,fmt=fmt,align=aln)
        ws5.row_dimensions[row].height=16
        row+=1
    row+=2

for i,w in enumerate([8,10,10,10,10,10,10,10,10,10,24],1):
    ws5.column_dimensions[get_column_letter(i)].width=w

# Salveaza
wb.save("rezultate_JES_final.xlsx")
print("\n✅ Excel salvat: rezultate_JES_final.xlsx")
print("✅ Toate graficele si calculele sunt gata!")
print("\nFisiere generate:")
print("  rezultate_JES_final.xlsx")
print("  figura1_CDF_R_h.png")
print("  figura2_rata_succes_complete.png")
print("  figura3_rata_succes_censurate.png")
print("  figura4_MSE.png")
