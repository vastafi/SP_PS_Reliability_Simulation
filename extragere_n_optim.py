"""
================================================================
Extragere n optim din fisierele CSV si calcule MLE

Formula: N = floor(((z*sigma)/epsilon)^2) + 1
epsilon = 0.1% din medie (eroare relativa)

Fisiere necesare:
  SP: 1 fisier  (435,781 retele >= n_SP=410,363)
  PS: 4 fisiere (1,743,124 retele >= n_PS=1,528,677)
================================================================
"""

import numpy as np
import math
from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════
# 1. PARAMETRI
# ════════════════════════════════════════════════════════════
K=9; N=3; M=2; alpha=0.05
z=norm.ppf(1-alpha/2)   # 1.9599639845

# ════════════════════════════════════════════════════════════
# 2. FUNCTII MODEL
# ════════════════════════════════════════════════════════════
def pmf_sp(k, K=9):
    pk=(k+1)/(K+1); pk0=k/(K+1)
    if k==0: return 1-(1-pk**N)**M
    return (1-pk0**N)**M-(1-pk**N)**M

def pmf_ps(k, K=9):
    qk=(K-k)/(K+1); qk0=(K-k+1)/(K+1)
    if k==0: return (1-qk**N)**M
    return (1-qk**N)**M-(1-qk0**N)**M

def F_sp(k, K=9): return 1-(1-((k+1)/(K+1))**N)**M
def F_ps(k, K=9): return (1-((K-k)/(K+1))**N)**M
def R_sp(k, K=9): return 1-F_sp(k,K)
def R_ps(k, K=9): return 1-F_ps(k,K)
def h_sp(k, K=9):
    p=pmf_sp(k,K); r=R_sp(k-1,K) if k>0 else 1.0
    return p/r if r>1e-15 else 0.0
def h_ps(k, K=9):
    p=pmf_ps(k,K); r=R_ps(k-1,K) if k>0 else 1.0
    return p/r if r>1e-15 else 0.0

mu_sp=sum(k*pmf_sp(k) for k in range(K+1))
mu_ps=sum(k*pmf_ps(k) for k in range(K+1))
var_sp=sum((k-mu_sp)**2*pmf_sp(k) for k in range(K+1))
var_ps=sum((k-mu_ps)**2*pmf_ps(k) for k in range(K+1))
sig=math.sqrt(var_sp)  # sig_sp = sig_ps prin simetrie

# ════════════════════════════════════════════════════════════
# 3. CALCULUL n PRIN FORMULA TLC (epsilon = 0.1% relativ)
# ════════════════════════════════════════════════════════════
eps_rel = 0.001  # 0.1%

eps_SP = eps_rel * mu_sp   # eroare absoluta pt SP
eps_PS = eps_rel * mu_ps   # eroare absoluta pt PS

n_SP = math.floor(((z * sig) / eps_SP)**2) + 1
n_PS = math.floor(((z * sig) / eps_PS)**2) + 1

print("="*65)
print("CALCULUL n — Formula: N = floor(((z*sigma)/epsilon)^2) + 1")
print("epsilon = 0.1% din medie (eroare relativa)")
print("="*65)
print(f"\nz_{{0.975}}  = {z:.10f}")
print(f"sigma      = {sig:.10f}  (sigma_SP = sigma_PS prin simetrie)")
print(f"E[U] SP    = {mu_sp:.10f}")
print(f"E[V] PS    = {mu_ps:.10f}")
print(f"\nPentru reteaua SP:")
print(f"  epsilon_SP = {eps_rel}*{mu_sp:.6f} = {eps_SP:.10f}")
print(f"  (z*sigma/eps_SP)^2 = {((z*sig)/eps_SP)**2:.4f}")
print(f"  n_SP = floor({((z*sig)/eps_SP)**2:.4f}) + 1 = {n_SP:,}")
print(f"\nPentru reteaua PS:")
print(f"  epsilon_PS = {eps_rel}*{mu_ps:.6f} = {eps_PS:.10f}")
print(f"  (z*sigma/eps_PS)^2 = {((z*sig)/eps_PS)**2:.4f}")
print(f"  n_PS = floor({((z*sig)/eps_PS)**2:.4f}) + 1 = {n_PS:,}")

# Cifre necesare
cifre_SP = n_SP * N * M
cifre_PS = n_PS * N * M
cifre_per_fisier = 2_614_690
retele_per_fisier = cifre_per_fisier // (N * M)  # 435,781

print(f"\nCifre necesare (N*M={N*M} cifre per retea):")
print(f"  SP: {n_SP:,} retele × {N*M} = {cifre_SP:,} cifre")
print(f"  PS: {n_PS:,} retele × {N*M} = {cifre_PS:,} cifre")
print(f"\nPer fisier: {cifre_per_fisier:,} cifre => {retele_per_fisier:,} retele")
print(f"\nFisiere necesare:")
print(f"  SP: ceil({n_SP:,}/{retele_per_fisier:,}) = {math.ceil(n_SP/retele_per_fisier)} fisier(e)")
print(f"  PS: ceil({n_PS:,}/{retele_per_fisier:,}) = {math.ceil(n_PS/retele_per_fisier)} fisier(e)")

# ════════════════════════════════════════════════════════════
# 4. FISIERELE CSV
# ════════════════════════════════════════════════════════════
BASE = "./"

FILES = {
    "Python secrets": [
        f"{BASE}data/python_secrets/secrets_data1.csv",
        f"{BASE}data/python_secrets/secrets_data2.csv",
        f"{BASE}data/python_secrets/secrets_data3.csv",
        f"{BASE}data/python_secrets/secrets_data4.csv",
        f"{BASE}data/python_secrets/secrets_data5.csv",
    ],
    "Cifrele lui pi": [
        f"{BASE}data/pi/pi_digits_part.csv",
        f"{BASE}data/pi/pi_digits_part1.csv",
        f"{BASE}data/pi/pi_digits_part2.csv",
        f"{BASE}data/pi/pi_digits_part3.csv",
        f"{BASE}data/pi/pi_digits_part4.csv",
        f"{BASE}data/pi/pi_digits_part5.csv",
    ],
    "Java ThreadLocalRandom": [
        f"{BASE}data/java_threadlocal/threadlocal_data1.csv",
        f"{BASE}data/java_threadlocal/threadlocal_data2.csv",
        f"{BASE}data/java_threadlocal/threadlocal_data3.csv",
        f"{BASE}data/java_threadlocal/threadlocal_data4.csv",
        f"{BASE}data/java_threadlocal/threadlocal_data5.csv",
    ],
}

# ════════════════════════════════════════════════════════════
# 5. FUNCTIE DE CITIRE — EXACT n_necesar RETELE
# ════════════════════════════════════════════════════════════

def citeste_retele(file_list, n_retele_necesar, net_type):
    """
    Citeste din fisierele CSV exact n_retele_necesar retele.
    Concateneaza fisierele daca e nevoie.
    Returneaza array de durate de viata ale retelei.
    """
    cifre_necesare = n_retele_necesar * N * M
    toate_cifrele = []

    for fpath in file_list:
        data = np.loadtxt(fpath, dtype=np.int32)
        toate_cifrele.append(data)
        if sum(len(x) for x in toate_cifrele) >= cifre_necesare:
            break

    # Concateneaza si taie exact
    cifre = np.concatenate(toate_cifrele)[:cifre_necesare]
    n_actual = len(cifre) // (N * M)

    # Reshape si calculeaza durata de viata
    mat = cifre[:n_actual * N * M].reshape(n_actual, M, N)

    if net_type == 'SP':
        # U = min_i(max_j(X_ij))
        return mat.max(axis=2).min(axis=1)
    else:
        # V = max_i(min_j(X_ij))
        return mat.min(axis=2).max(axis=1)

# ════════════════════════════════════════════════════════════
# 6. MLE
# ════════════════════════════════════════════════════════════

def mle_complete(data, pmf_func, K_max=30):
    Km=int(np.max(data)); best_K,best_ll=Km,-np.inf
    vals,cnts=np.unique(data,return_counts=True)
    for Kt in range(Km, K_max+1):
        ll=sum(cnt*np.log(max(pmf_func(int(v),Kt),1e-300))
               for v,cnt in zip(vals,cnts))
        if ll>best_ll: best_ll,best_K=ll,Kt
    return best_K

def mle_censored(data, intervals, F_func, K_max=30):
    counts=[]
    for a,b in intervals:
        if b is None: counts.append(int(np.sum(data>=a)))
        else: counts.append(int(np.sum((data>=a)&(data<b))))
    Km=int(intervals[-1][0]); best_K,best_ll=Km,-np.inf
    for Kt in range(Km, K_max+1):
        ll=0; ok=True
        for (a,b),cnt in zip(intervals,counts):
            if cnt==0: continue
            Fb=F_func(b-1,Kt) if b is not None else 1.0
            Fa=F_func(a-1,Kt) if a>0 else 0.0
            p=Fb-Fa
            if p<=0: ok=False; break
            ll+=cnt*np.log(p)
        if ok and ll>best_ll: best_ll,best_K=ll,Kt
    return best_K

INT_SP_C3=[(0,4),(4,7),(7,None)]
INT_SP_C5=[(0,4),(4,6),(6,7),(7,8),(8,None)]
INT_PS_C3=[(0,3),(3,6),(6,None)]
INT_PS_C5=[(0,3),(3,5),(5,7),(7,8),(8,None)]

# ════════════════════════════════════════════════════════════
# 7. CALCULE PRINCIPALE
# ════════════════════════════════════════════════════════════

print("\n" + "="*65)
print("CALCULE MLE PE DATE REALE")
print("="*65)

GNPA_ORDER = ["Java ThreadLocalRandom", "Python secrets", "Cifrele lui pi"]
results = {}

for gname in GNPA_ORDER:
    results[gname] = {}
    print(f"\n{'─'*65}")
    print(f"GNPA: {gname}")
    print(f"{'─'*65}")

    for net, pmf_f, F_f, ic3, ic5, n_nec, mu_t in [
        ('SP', pmf_sp, F_sp, INT_SP_C3, INT_SP_C5, n_SP, mu_sp),
        ('PS', pmf_ps, F_ps, INT_PS_C3, INT_PS_C5, n_PS, mu_ps),
    ]:
        print(f"\n  Reteaua {net}: n={n_nec:,} retele necesare")

        lifetimes = citeste_retele(FILES[gname], n_nec, net)
        n_actual = len(lifetimes)

        mean_sim = float(np.mean(lifetimes))
        Kc  = mle_complete(lifetimes, pmf_f)
        Kc3 = mle_censored(lifetimes, ic3, F_f)
        Kc5 = mle_censored(lifetimes, ic5, F_f)

        print(f"    n_actual   = {n_actual:,}")
        print(f"    mean_sim   = {mean_sim:.10f}")
        print(f"    E[X]_teor  = {mu_t:.10f}")
        print(f"    |mean-teor|= {abs(mean_sim-mu_t):.10f}  "
              f"(< epsilon={0.001*mu_t:.6f} ✅)" if abs(mean_sim-mu_t) < 0.001*mu_t
              else f"    |mean-teor|= {abs(mean_sim-mu_t):.10f}")
        print(f"    K_hat_C    = {Kc}    (K_adevarat = {K})")
        print(f"    K_hat_C3   = {Kc3}")
        print(f"    K_hat_C5   = {Kc5}")

        results[gname][net] = {
            'n': n_actual, 'n_necesar': n_nec,
            'mean_sim': mean_sim, 'mu_teoretic': mu_t,
            'Kc': Kc, 'Kc3': Kc3, 'Kc5': Kc5,
            'mse_c':  float((Kc-K)**2),
            'mse_c3': float((Kc3-K)**2),
            'mse_c5': float((Kc5-K)**2),
        }

# ════════════════════════════════════════════════════════════
# 8. TABEL COMPARATIV
# ════════════════════════════════════════════════════════════
print("\n" + "="*65)
print("TABEL COMPARATIV FINAL")
print("="*65)
for net in ['SP','PS']:
    n_nec = n_SP if net=='SP' else n_PS
    mu_t  = mu_sp if net=='SP' else mu_ps
    print(f"\nReteaua {net} (n={n_nec:,}, E[X]={mu_t:.6f}):")
    print(f"  {'GNPA':<26} {'Mean_sim':>14} {'K_c':>6} {'K_c3':>6} {'K_c5':>6} {'MSE_c':>8} {'MSE_c3':>8} {'MSE_c5':>8}")
    print(f"  {'-'*80}")
    for gname in GNPA_ORDER:
        res=results[gname][net]
        print(f"  {gname:<26} {res['mean_sim']:>14.10f} "
              f"{res['Kc']:>6} {res['Kc3']:>6} {res['Kc5']:>6} "
              f"{res['mse_c']:>8.4f} {res['mse_c3']:>8.4f} {res['mse_c5']:>8.4f}")

# ════════════════════════════════════════════════════════════
# 9. SALVARE EXCEL
# ════════════════════════════════════════════════════════════
wb=Workbook()
HDR="1F4E79"; SUB="2E75B6"; GRN="E2EFDA"; YEL="FFF2CC"; RED="FCE4D6"

def cell(ws,r,c,v,bold=False,bg=None,fmt=None,
         align="center",color="000000",italic=False,size=10,fname="Arial"):
    cl=ws.cell(row=r,column=c,value=v)
    cl.font=Font(name=fname,bold=bold,size=size,color=color,italic=italic)
    if bg: cl.fill=PatternFill("solid",fgColor=bg)
    cl.alignment=Alignment(horizontal=align,vertical="center",wrap_text=True)
    t=Side(style="thin",color="AAAAAA")
    cl.border=Border(left=t,right=t,top=t,bottom=t)
    if fmt: cl.number_format=fmt
    return cl

NF10="0.0000000000"; NF6="0.000000"; NF4="0.0000"

# ── F1: Calculul n ────────────────────────────────────────────
ws1=wb.active; ws1.title="Calculul_n"
ws1.merge_cells("A1:F1")
cell(ws1,1,1,
     "CALCULUL n — N=floor(((z*sigma)/epsilon)^2)+1  |  epsilon=0.1% din medie",
     bold=True,bg=HDR,color="FFFFFF",size=12)
ws1.row_dimensions[1].height=28

# Parametri
ws1.merge_cells("A3:F3")
cell(ws1,3,1,"PARAMETRI",bold=True,bg=SUB,color="FFFFFF")
params=[
    ("z_{0.975} (alpha=0.05)", z, NF10),
    ("sigma_SP = sigma_PS",   sig, NF10),
    ("E[U] reteaua SP",       mu_sp, NF10),
    ("E[V] reteaua PS",       mu_ps, NF10),
    ("epsilon_SP = 0.001*E[U]", eps_SP, NF10),
    ("epsilon_PS = 0.001*E[V]", eps_PS, NF10),
    ("n_SP = floor(((z*sig)/eps_SP)^2)+1", n_SP, "0"),
    ("n_PS = floor(((z*sig)/eps_PS)^2)+1", n_PS, "0"),
    ("Cifre per retea (N*M)", N*M, "0"),
    ("Cifre per fisier CSV", 2_614_690, "0"),
    ("Retele per fisier", retele_per_fisier, "0"),
    ("Fisiere necesare SP", math.ceil(n_SP/retele_per_fisier), "0"),
    ("Fisiere necesare PS", math.ceil(n_PS/retele_per_fisier), "0"),
]
for i,(nm,vl,fmt) in enumerate(params):
    r=4+i; bg=GRN if i%2==0 else YEL
    cell(ws1,r,1,nm,bg=bg,align="left")
    ws1.merge_cells(f"B{r}:C{r}")
    cell(ws1,r,2,vl,bg=bg,fmt=fmt,align="right",
         fname="Courier New",bold=(i>=6))
    ws1.row_dimensions[r].height=16

# Tabel detaliat
ws1.merge_cells("A19:F19")
cell(ws1,19,1,"DETALII CALCUL (10 zecimale)",bold=True,bg=SUB,color="FFFFFF")
for j,h in enumerate(["Reteaua","E[X]","epsilon=0.001*E[X]",
                       "(z*sig/eps)^2","floor(...)","N=floor+1"],1):
    cell(ws1,20,j,h,bold=True,bg=HDR,color="FFFFFF")
ws1.row_dimensions[20].height=20
for i,(net,mu,eps,n) in enumerate([
    ("SP",mu_sp,eps_SP,n_SP),
    ("PS",mu_ps,eps_PS,n_PS)]):
    r=21+i; bg=GRN if i==0 else YEL
    ratio=((z*sig)/eps)**2; fl=math.floor(ratio)
    for j,v in enumerate([net,mu,eps,ratio,fl,n],1):
        cell(ws1,r,j,v,bg=bg,
             fmt=None if j in {1,5,6} else NF10,
             align="left" if j==1 else ("right" if j>1 else "center"),
             bold=(j==6))
    ws1.row_dimensions[r].height=16

for i,w in enumerate([30,18,20,22,14,16],1):
    ws1.column_dimensions[get_column_letter(i)].width=w

# ── F2: Valori Teoretice ──────────────────────────────────────
ws2=wb.create_sheet("Valori_Teoretice")
ws2.merge_cells("A1:I1")
cell(ws2,1,1,"VALORI TEORETICE — N=3, M=2, K=9 (10 zecimale)",
     bold=True,bg=HDR,color="FFFFFF",size=12)
ws2.row_dimensions[1].height=28
for j,h in enumerate(["k","F_sp(k)","P(U=k)","R_sp(k)","h_sp(k)",
                           "F_ps(k)","P(V=k)","R_ps(k)","h_ps(k)"],1):
    cell(ws2,3,j,h,bold=True,bg=HDR,color="FFFFFF")
ws2.row_dimensions[3].height=20
for k in range(K+1):
    r=4+k; bg=GRN if k%2==0 else None
    for j,v in enumerate([k,F_sp(k),pmf_sp(k),R_sp(k),h_sp(k),
                              F_ps(k),pmf_ps(k),R_ps(k),h_ps(k)],1):
        cell(ws2,r,j,v,bg=bg,
             fmt=None if j==1 else NF10,
             align="center" if j==1 else "right")
    ws2.row_dimensions[r].height=16
r_s=4+K+1
cell(ws2,r_s,1,"SUMA PMF",bold=True,bg=YEL)
cell(ws2,r_s,3,sum(pmf_sp(k) for k in range(K+1)),
     bold=True,bg=YEL,fmt=NF10,align="right")
cell(ws2,r_s,7,sum(pmf_ps(k) for k in range(K+1)),
     bold=True,bg=YEL,fmt=NF10,align="right")
for i,w in enumerate([6,14,14,14,14,14,14,14,14],1):
    ws2.column_dimensions[get_column_letter(i)].width=w

# ── F3: Rezultate MLE ────────────────────────────────────────
ws3=wb.create_sheet("Rezultate_MLE")
ws3.merge_cells("A1:K1")
cell(ws3,1,1,
     f"REZULTATE MLE — Date reale CSV | N={N}, M={M}, K={K} | R=5 per GNPA",
     bold=True,bg=HDR,color="FFFFFF",size=12)
ws3.row_dimensions[1].height=28
ws3.merge_cells("A3:K3")
cell(ws3,3,1,
     f"n_SP={n_SP:,} retele (eps={eps_SP:.6f}) | "
     f"n_PS={n_PS:,} retele (eps={eps_PS:.6f}) | "
     f"z={z:.6f} | sigma={sig:.6f}",
     bg="DEEAF1",align="left",italic=True,size=9)

for j,h in enumerate(["Reteaua","GNPA","n_necesar","n_actual",
                       "Mean_sim","E[X]_teoretic","|diff|",
                       "K_hat C","K_hat C3","K_hat C5",
                       "Concluzie"],1):
    cell(ws3,5,j,h,bold=True,bg=HDR,color="FFFFFF")
ws3.row_dimensions[5].height=35

row=6
for net,n_nec,mu_t in [('SP',n_SP,mu_sp),('PS',n_PS,mu_ps)]:
    for gi,gname in enumerate(GNPA_ORDER):
        bg=GRN if gi==0 else (YEL if gi==1 else None)
        res=results[gname][net]
        diff=abs(res['mean_sim']-mu_t)
        ok=diff<0.001*mu_t
        concl="K=9" if res['Kc']==K and res['Kc3']==K and res['Kc5']==K \
              else f"K_c={res['Kc']},K_c3={res['Kc3']},K_c5={res['Kc5']}"
        vals=[net if gi==0 else "",gname,n_nec,res['n'],
              res['mean_sim'],mu_t,diff,
              res['Kc'],res['Kc3'],res['Kc5'],concl]
        fmts=[None,None,"0","0",NF10,NF10,NF10,"0","0","0",None]
        aligns=["center","left","right","right",
                "right","right","right",
                "center","center","center","left"]
        for j,(v,fmt,aln) in enumerate(zip(vals,fmts,aligns),1):
            cell(ws3,row,j,v,bg=bg,fmt=fmt,align=aln,
                 bold=(j==7 and not ok),
                 color="C00000" if (j==7 and not ok) else "000000")
        ws3.row_dimensions[row].height=16
        row+=1
    row+=1

for i,w in enumerate([8,26,12,12,16,16,16,10,10,10,20],1):
    ws3.column_dimensions[get_column_letter(i)].width=w

wb.save("rezultate_n_optim.xlsx")
print("\nExcel salvat: rezultate_n_optim.xlsx")
print("\nFisier generat: rezultate_n_optim.xlsx")
